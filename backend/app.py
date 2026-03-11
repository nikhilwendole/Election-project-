"""
app.py — Flask REST API for Election Face Detection System
"""

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import cv2
import numpy as np
import base64
import os
import io
from datetime import datetime
from dotenv import load_dotenv
from pymongo import MongoClient
import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

load_dotenv()

app = Flask(__name__)
CORS(app)

# ── MongoDB ──────────────────────────────────────────────────
MONGO_URI       = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
DB_NAME         = os.getenv("DB_NAME", "election_system")
COLLECTION_NAME = os.getenv("COLLECTION_NAME", "voters")

client     = MongoClient(MONGO_URI)
db         = client[DB_NAME]
voters_col = db[COLLECTION_NAME]
admins_col = db["admins"]

# ── Default Admin ────────────────────────────────────────────
if admins_col.count_documents({}) == 0:
    admins_col.insert_one({
        "username"  : "admin",
        "password"  : hashlib.sha256("admin123".encode()).hexdigest(),
        "created_at": datetime.utcnow()
    })
    print("[DB] Default admin created — username: admin | password: admin123")

# ── Face Detection ───────────────────────────────────────────
FACE_CASCADE = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)
EYE_CASCADE = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_eye.xml"
)

THRESHOLD = 0.92


# ════════════════════════════════════════════════════════════
#  FACE FUNCTIONS
# ════════════════════════════════════════════════════════════

def decode_image(base64_str):
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
    img_bytes = base64.b64decode(base64_str)
    img_array = np.frombuffer(img_bytes, dtype=np.uint8)
    return cv2.imdecode(img_array, cv2.IMREAD_COLOR)


def get_facial_geometry(gray, face_coords):
    x, y, w, h = face_coords
    pad = 15
    y1  = max(0, y - pad)
    y2  = min(gray.shape[0], y + h + pad)
    x1  = max(0, x - pad)
    x2  = min(gray.shape[1], x + w + pad)

    face = gray[y1:y2, x1:x2]
    face = cv2.resize(face, (200, 200))
    face = cv2.equalizeHist(face)

    features = []

    # 1. Zone analysis 5x5 grid
    zone_h = face.shape[0] // 5
    zone_w = face.shape[1] // 5
    for row in range(5):
        for col in range(5):
            zone = face[row*zone_h:(row+1)*zone_h,
                        col*zone_w:(col+1)*zone_w]
            features.append(float(np.mean(zone)))
            features.append(float(np.std(zone)))

    # 2. Facial proportions
    features.append(float(w) / (float(h) + 1e-5))
    features.append(float(np.mean(face[:100, :])) /
                   (float(np.mean(face[100:, :])) + 1e-5))
    features.append(float(np.mean(face[:, :100])) /
                   (float(np.mean(face[:, 100:])) + 1e-5))

    # 3. Eye geometry
    eyes = EYE_CASCADE.detectMultiScale(face, 1.1, 5, minSize=(20, 20))
    if len(eyes) >= 2:
        eyes    = sorted(eyes, key=lambda e: e[0])
        ex1, ey1, ew1, eh1 = eyes[0]
        ex2, ey2, ew2, eh2 = eyes[1]
        eye1_cx = (ex1 + ew1/2) / 200.0
        eye1_cy = (ey1 + eh1/2) / 200.0
        eye2_cx = (ex2 + ew2/2) / 200.0
        eye2_cy = (ey2 + eh2/2) / 200.0
        eye_dist = np.sqrt((eye2_cx-eye1_cx)**2 + (eye2_cy-eye1_cy)**2)
        features.extend([eye1_cx, eye1_cy, eye2_cx, eye2_cy,
                         float(eye_dist), ew1/(ew2+1e-5),
                         abs(eye1_cy-eye2_cy)])
    else:
        features.extend([0.0] * 7)

    # 4. Gradient features
    sobelx    = cv2.Sobel(face, cv2.CV_64F, 1, 0, ksize=3)
    sobely    = cv2.Sobel(face, cv2.CV_64F, 0, 1, ksize=3)
    magnitude = np.sqrt(sobelx**2 + sobely**2)
    bh = magnitude.shape[0] // 4
    bw = magnitude.shape[1] // 4
    for row in range(4):
        for col in range(4):
            block = magnitude[row*bh:(row+1)*bh, col*bw:(col+1)*bw]
            features.append(float(np.mean(block)))
            features.append(float(np.std(block)))

    vec  = np.array(features, dtype=np.float32)
    norm = np.linalg.norm(vec)
    if norm > 0:
        vec = vec / norm
    return vec.tolist()


def extract_encoding(frame):
    gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = FACE_CASCADE.detectMultiScale(
        gray, scaleFactor=1.05, minNeighbors=6, minSize=(100, 100)
    )
    if len(faces) == 0:
        return None, "no_face"
    if len(faces) > 1:
        return None, "multiple_faces"
    encoding = get_facial_geometry(gray, faces[0])
    return encoding, "ok"


def compare_encodings(enc1, enc2):
    a = np.array(enc1, dtype=np.float32)
    b = np.array(enc2, dtype=np.float32)
    if a.shape != b.shape:
        return 0.0
    dot = np.dot(a, b)
    na  = np.linalg.norm(a)
    nb  = np.linalg.norm(b)
    if na == 0 or nb == 0:
        return 0.0
    return float(dot / (na * nb))


def check_duplicate(new_encoding):
    voters = list(voters_col.find({}, {"_id": 0}))
    if not voters:
        return False, None, 0

    best_score = -1
    best_voter = None

    for voter in voters:
        stored = voter.get("face_encoding", [])
        if len(stored) != len(new_encoding):
            continue
        score = compare_encodings(new_encoding, stored)
        if score > best_score:
            best_score = score
            best_voter = voter

    print(f"[DEBUG] Best match: {round(best_score*100, 2)}%")

    if best_score >= THRESHOLD:
        return True, best_voter, round(best_score * 100, 2)
    return False, None, 0


def generate_voter_id(encoding):
    raw = str(np.round(encoding, 4)).encode()
    return hashlib.sha256(raw).hexdigest()[:20]


# ════════════════════════════════════════════════════════════
#  API ROUTES
# ════════════════════════════════════════════════════════════

@app.route("/api/login", methods=["POST"])
def login():
    data     = request.json
    username = data.get("username", "")
    password = hashlib.sha256(
        data.get("password", "").encode()
    ).hexdigest()
    admin = admins_col.find_one({"username": username, "password": password})
    if admin:
        return jsonify({"success": True, "message": "Login successful"})
    return jsonify({"success": False, "message": "Invalid credentials"}), 401


@app.route("/api/verify-face", methods=["POST"])
def verify_face():
    data      = request.json
    image_b64 = data.get("image")
    if not image_b64:
        return jsonify({"status": "error", "message": "No image"}), 400

    frame             = decode_image(image_b64)
    encoding, status  = extract_encoding(frame)

    if status == "no_face":
        return jsonify({"status": "no_face", "message": "No face detected"})
    if status == "multiple_faces":
        return jsonify({"status": "multiple_faces", "message": "Multiple faces detected"})

    is_dup, matched, confidence = check_duplicate(encoding)

    if is_dup:
        return jsonify({
            "status"    : "duplicate",
            "message"   : "Already voted!",
            "confidence": confidence,
            "encoding"  : encoding,
            "voter"     : {
                "name"        : matched["voter_name"],
                "voter_id"    : matched["voter_id"],
                "voter_number": matched.get("voter_number", ""),
                "voted_at"    : str(matched.get("voted_at", "")),
            }
        })

    return jsonify({
        "status"  : "new",
        "message" : "New voter verified!",
        "encoding": encoding
    })


@app.route("/api/cast-vote", methods=["POST"])
def cast_vote():
    data         = request.json
    name         = data.get("name", "").strip()
    voter_number = data.get("voter_number", "").strip()
    encoding     = data.get("encoding", [])

    if not name or not encoding:
        return jsonify({"success": False, "message": "Name and encoding required"}), 400

    voter_id = generate_voter_id(encoding)
    document = {
        "voter_id"     : voter_id,
        "voter_name"   : name,
        "voter_number" : voter_number or f"VN-{voter_id[:8].upper()}",
        "face_encoding": encoding,
        "voted_at"     : datetime.utcnow(),
        "status"       : "voted"
    }
    voters_col.insert_one(document)

    return jsonify({
        "success"     : True,
        "message"     : "Vote cast successfully!",
        "voter_id"    : voter_id,
        "voter_name"  : name,
        "voter_number": document["voter_number"],
        "voted_at"    : str(document["voted_at"])
    })


@app.route("/api/stats", methods=["GET"])
def get_stats():
    total = voters_col.count_documents({})
    today = datetime.utcnow().replace(hour=0, minute=0, second=0)
    today_count = voters_col.count_documents({"voted_at": {"$gte": today}})

    pipeline = [
        {"$group": {"_id": {"$hour": "$voted_at"}, "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    hourly      = list(voters_col.aggregate(pipeline))
    hourly_data = [{"hour": f"{h['_id']}:00", "votes": h["count"]} for h in hourly]

    return jsonify({
        "total_voters": total,
        "today_votes" : today_count,
        "hourly_data" : hourly_data,
    })


@app.route("/api/voters", methods=["GET"])
def get_voters():
    voters = list(voters_col.find({}, {"_id": 0, "face_encoding": 0}))
    for v in voters:
        v["voted_at"] = str(v.get("voted_at", ""))
    return jsonify({"voters": voters, "total": len(voters)})


@app.route("/api/voters/<voter_id>", methods=["DELETE"])
def delete_voter(voter_id):
    result = voters_col.delete_one({"voter_id": voter_id})
    if result.deleted_count:
        return jsonify({"success": True, "message": "Voter deleted"})
    return jsonify({"success": False, "message": "Voter not found"}), 404


@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    voters = list(voters_col.find({}, {"_id": 0, "face_encoding": 0}))
    wb     = openpyxl.Workbook()
    ws     = wb.active
    ws.title = "Voters"

    headers = ["#", "Name", "Voter Number", "Voter ID", "Voted At", "Status"]
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1a56db")
        cell.alignment = Alignment(horizontal="center")

    for i, voter in enumerate(voters, 1):
        ws.append([
            i,
            voter.get("voter_name", ""),
            voter.get("voter_number", ""),
            voter.get("voter_id", ""),
            str(voter.get("voted_at", "")),
            voter.get("status", "voted")
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="voters.xlsx")


@app.route("/api/export/pdf", methods=["GET"])
def export_pdf():
    voters = list(voters_col.find({}, {"_id": 0, "face_encoding": 0}))
    output = io.BytesIO()
    doc    = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    story  = []

    story.append(Paragraph("Election Voters Report", styles["Title"]))
    story.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC",
        styles["Normal"]
    ))

    table_data = [["#", "Name", "Voter Number", "Voter ID", "Voted At"]]
    for i, voter in enumerate(voters, 1):
        table_data.append([
            str(i),
            voter.get("voter_name", ""),
            voter.get("voter_number", ""),
            voter.get("voter_id", "")[:12] + "...",
            str(voter.get("voted_at", ""))[:16]
        ])

    table = Table(table_data, colWidths=[30, 120, 100, 110, 120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a56db")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f3f4f6")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
    ]))
    story.append(table)
    doc.build(story)

    output.seek(0)
    return send_file(output,
                     mimetype="application/pdf",
                     as_attachment=True,
                     download_name="voters.pdf")


# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n╔══════════════════════════════════════════╗")
    print("║   Election API Server Starting...        ║")
    print("║   http://localhost:5000                  ║")
    print("╚══════════════════════════════════════════╝\n")
    app.run(debug=True, port=5000)
















































# """
# app.py — Flask REST API for Election Face Detection System
# With Anti-Spoofing / Liveness Detection
# """

# from flask import Flask, jsonify, request, send_file
# from flask_cors import CORS
# import cv2
# import numpy as np
# import base64
# import os
# import io
# from datetime import datetime
# from dotenv import load_dotenv
# from pymongo import MongoClient
# import hashlib
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment
# from reportlab.lib.pagesizes import letter
# from reportlab.lib import colors
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
# from reportlab.lib.styles import getSampleStyleSheet

# load_dotenv()

# app = Flask(__name__)
# CORS(app)

# # ── MongoDB ──────────────────────────────────────────────────
# MONGO_URI       = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
# DB_NAME         = os.getenv("DB_NAME", "election_system")
# COLLECTION_NAME = os.getenv("COLLECTION_NAME", "voters")

# client     = MongoClient(MONGO_URI)
# db         = client[DB_NAME]
# voters_col = db[COLLECTION_NAME]
# admins_col = db["admins"]

# # ── Default Admin ────────────────────────────────────────────
# if admins_col.count_documents({}) == 0:
#     admins_col.insert_one({
#         "username"  : "admin",
#         "password"  : hashlib.sha256("admin123".encode()).hexdigest(),
#         "created_at": datetime.utcnow()
#     })
#     print("[DB] Default admin created — username: admin | password: admin123")

# # ── Face Detection Setup ─────────────────────────────────────
# FACE_CASCADE = cv2.CascadeClassifier(
#     cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
# )
# EYE_CASCADE = cv2.CascadeClassifier(
#     cv2.data.haarcascades + "haarcascade_eye.xml"
# )

# THRESHOLD = 0.92


# # ════════════════════════════════════════════════════════════
# #  HELPER FUNCTIONS
# # ════════════════════════════════════════════════════════════

# def decode_image(base64_str):
#     """Convert base64 image from React webcam to OpenCV frame."""
#     if "," in base64_str:
#         base64_str = base64_str.split(",")[1]
#     img_bytes = base64.b64decode(base64_str)
#     img_array = np.frombuffer(img_bytes, dtype=np.uint8)
#     return cv2.imdecode(img_array, cv2.IMREAD_COLOR)


# def check_liveness(frame):
#     """
#     Detect if face is REAL or a printed photo / screen image.
#     Tests: texture variance, color channels, gradient diversity, frequency analysis
#     Returns (is_live, reason)
#     """
#     gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
#     faces = FACE_CASCADE.detectMultiScale(
#         gray, scaleFactor=1.05, minNeighbors=6, minSize=(100, 100)
#     )

#     if len(faces) == 0:
#         return False, "no_face"
#     if len(faces) > 1:
#         return False, "multiple_faces"

#     x, y, w, h = faces[0]

#     # Crop face with padding
#     pad        = 20
#     y1         = max(0, y - pad)
#     y2         = min(frame.shape[0], y + h + pad)
#     x1         = max(0, x - pad)
#     x2         = min(frame.shape[1], x + w + pad)
#     face_gray  = cv2.resize(gray[y1:y2, x1:x2],  (100, 100))
#     face_color = cv2.resize(frame[y1:y2, x1:x2], (100, 100))

#     # ── Test 1: Texture variance ─────────────────────────────
#     # Real faces have high texture variance
#     # Printed photos are flat and smooth
#     laplacian = cv2.Laplacian(face_gray, cv2.CV_64F)
#     variance  = laplacian.var()
#     print(f"[LIVENESS] Texture variance : {variance:.2f}  (min: 80)")
#     if variance < 80:
#         return False, "spoof_photo"

#     # ── Test 2: Color channel analysis ───────────────────────
#     # Real skin: R > G > B
#     # Printed photo: channels are nearly equal
#     b, g, r  = cv2.split(face_color)
#     r_mean   = float(np.mean(r))
#     b_mean   = float(np.mean(b))
#     color_diff = abs(r_mean - b_mean)
#     print(f"[LIVENESS] Color diff R-B  : {color_diff:.2f}  (min: 5)")
#     if color_diff < 5:
#         return False, "spoof_photo"

#     # ── Test 3: Gradient diversity ───────────────────────────
#     # Real faces have diverse edge gradients
#     # Photos have uniform/flat gradients
#     sobelx     = cv2.Sobel(face_gray, cv2.CV_64F, 1, 0, ksize=3)
#     sobely     = cv2.Sobel(face_gray, cv2.CV_64F, 0, 1, ksize=3)
#     grad_x_std = np.std(sobelx)
#     grad_y_std = np.std(sobely)
#     print(f"[LIVENESS] Gradient X:{grad_x_std:.2f} Y:{grad_y_std:.2f}  (min: 15)")
#     if grad_x_std < 15 or grad_y_std < 15:
#         return False, "spoof_photo"

#     # ── Test 4: High frequency detail ───────────────────────
#     # Real faces have natural micro-texture (pores, skin detail)
#     # Photos are smoother at high frequencies
#     dft           = np.fft.fft2(face_gray)
#     dft_shift     = np.fft.fftshift(dft)
#     magnitude     = np.abs(dft_shift)
#     h2, w2        = magnitude.shape
#     center        = magnitude[h2//4:3*h2//4, w2//4:3*w2//4]
#     outer         = magnitude.copy()
#     outer[h2//4:3*h2//4, w2//4:3*w2//4] = 0
#     hf_ratio      = np.mean(outer) / (np.mean(center) + 1e-5)
#     print(f"[LIVENESS] HF ratio        : {hf_ratio:.4f}  (min: 0.01)")
#     if hf_ratio < 0.01:
#         return False, "spoof_photo"

#     print("[LIVENESS] ✔ Face is LIVE")
#     return True, "live"


# def get_facial_geometry(gray, face_coords):
#     """Extract unique facial geometry measurements as feature vector."""
#     x, y, w, h = face_coords
#     pad = 15
#     y1  = max(0, y - pad)
#     y2  = min(gray.shape[0], y + h + pad)
#     x1  = max(0, x - pad)
#     x2  = min(gray.shape[1], x + w + pad)

#     face = cv2.resize(gray[y1:y2, x1:x2], (200, 200))
#     face = cv2.equalizeHist(face)

#     features = []

#     # 1. Zone analysis — 5x5 grid
#     zone_h = face.shape[0] // 5
#     zone_w = face.shape[1] // 5
#     for row in range(5):
#         for col in range(5):
#             zone = face[row*zone_h:(row+1)*zone_h,
#                         col*zone_w:(col+1)*zone_w]
#             features.append(float(np.mean(zone)))
#             features.append(float(np.std(zone)))

#     # 2. Facial proportions
#     features.append(float(w) / (float(h) + 1e-5))
#     features.append(float(np.mean(face[:100, :])) /
#                    (float(np.mean(face[100:, :])) + 1e-5))
#     features.append(float(np.mean(face[:, :100])) /
#                    (float(np.mean(face[:, 100:])) + 1e-5))

#     # 3. Eye geometry
#     eyes = EYE_CASCADE.detectMultiScale(face, 1.1, 5, minSize=(20, 20))
#     if len(eyes) >= 2:
#         eyes    = sorted(eyes, key=lambda e: e[0])
#         ex1, ey1, ew1, eh1 = eyes[0]
#         ex2, ey2, ew2, eh2 = eyes[1]
#         eye1_cx  = (ex1 + ew1/2) / 200.0
#         eye1_cy  = (ey1 + eh1/2) / 200.0
#         eye2_cx  = (ex2 + ew2/2) / 200.0
#         eye2_cy  = (ey2 + eh2/2) / 200.0
#         eye_dist = np.sqrt((eye2_cx-eye1_cx)**2 + (eye2_cy-eye1_cy)**2)
#         features.extend([eye1_cx, eye1_cy, eye2_cx, eye2_cy,
#                          float(eye_dist), ew1/(ew2+1e-5),
#                          abs(eye1_cy-eye2_cy)])
#     else:
#         features.extend([0.0] * 7)

#     # 4. Gradient features — 4x4 blocks
#     sobelx    = cv2.Sobel(face, cv2.CV_64F, 1, 0, ksize=3)
#     sobely    = cv2.Sobel(face, cv2.CV_64F, 0, 1, ksize=3)
#     magnitude = np.sqrt(sobelx**2 + sobely**2)
#     bh = magnitude.shape[0] // 4
#     bw = magnitude.shape[1] // 4
#     for row in range(4):
#         for col in range(4):
#             block = magnitude[row*bh:(row+1)*bh, col*bw:(col+1)*bw]
#             features.append(float(np.mean(block)))
#             features.append(float(np.std(block)))

#     vec  = np.array(features, dtype=np.float32)
#     norm = np.linalg.norm(vec)
#     if norm > 0:
#         vec = vec / norm
#     return vec.tolist()


# def extract_encoding(frame):
#     """
#     Full pipeline:
#     1. Detect face
#     2. Check liveness (anti-spoof)
#     3. Extract face encoding
#     """
#     gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
#     faces = FACE_CASCADE.detectMultiScale(
#         gray, scaleFactor=1.05, minNeighbors=6, minSize=(100, 100)
#     )
#     if len(faces) == 0:
#         return None, "no_face"
#     if len(faces) > 1:
#         return None, "multiple_faces"

#     # Anti-spoofing check
#     is_live, reason = check_liveness(frame)
#     if not is_live:
#         if reason == "spoof_photo":
#             return None, "spoof"
#         return None, reason

#     encoding = get_facial_geometry(gray, faces[0])
#     return encoding, "ok"


# def compare_encodings(enc1, enc2):
#     """Cosine similarity between two face encodings."""
#     a = np.array(enc1, dtype=np.float32)
#     b = np.array(enc2, dtype=np.float32)
#     if a.shape != b.shape:
#         return 0.0
#     dot = np.dot(a, b)
#     na  = np.linalg.norm(a)
#     nb  = np.linalg.norm(b)
#     if na == 0 or nb == 0:
#         return 0.0
#     return float(dot / (na * nb))


# def check_duplicate(new_encoding):
#     """Check if encoding matches any existing voter in MongoDB."""
#     voters = list(voters_col.find({}, {"_id": 0}))
#     if not voters:
#         return False, None, 0

#     best_score = -1
#     best_voter = None

#     for voter in voters:
#         stored = voter.get("face_encoding", [])
#         if len(stored) != len(new_encoding):
#             continue
#         score = compare_encodings(new_encoding, stored)
#         if score > best_score:
#             best_score = score
#             best_voter = voter

#     print(f"[DUPLICATE] Best match: {round(best_score*100, 2)}%  (threshold: {THRESHOLD*100}%)")

#     if best_score >= THRESHOLD:
#         return True, best_voter, round(best_score * 100, 2)
#     return False, None, 0


# def generate_voter_id(encoding):
#     """Generate unique voter ID from face encoding."""
#     raw = str(np.round(encoding, 4)).encode()
#     return hashlib.sha256(raw).hexdigest()[:20]


# # ════════════════════════════════════════════════════════════
# #  API ROUTES
# # ════════════════════════════════════════════════════════════

# @app.route("/api/login", methods=["POST"])
# def login():
#     data     = request.json
#     username = data.get("username", "")
#     password = hashlib.sha256(data.get("password", "").encode()).hexdigest()
#     admin    = admins_col.find_one({"username": username, "password": password})
#     if admin:
#         return jsonify({"success": True, "message": "Login successful"})
#     return jsonify({"success": False, "message": "Invalid credentials"}), 401


# @app.route("/api/verify-face", methods=["POST"])
# def verify_face():
#     """
#     Main verification endpoint.
#     Returns: new / duplicate / no_face / multiple_faces / spoof
#     """
#     data      = request.json
#     image_b64 = data.get("image")
#     if not image_b64:
#         return jsonify({"status": "error", "message": "No image provided"}), 400

#     frame            = decode_image(image_b64)
#     encoding, status = extract_encoding(frame)

#     # ── Status responses ─────────────────────────────────────
#     if status == "no_face":
#         return jsonify({
#             "status" : "no_face",
#             "message": "No face detected. Please center yourself."
#         })

#     if status == "multiple_faces":
#         return jsonify({
#             "status" : "multiple_faces",
#             "message": "Multiple faces detected. Only one voter allowed."
#         })

#     if status == "spoof":
#         return jsonify({
#             "status" : "spoof",
#             "message": "Fake face detected! Please use your real face."
#         })

#     # ── Duplicate check ──────────────────────────────────────
#     is_dup, matched, confidence = check_duplicate(encoding)

#     if is_dup:
#         return jsonify({
#             "status"    : "duplicate",
#             "message"   : "This voter has already voted!",
#             "confidence": confidence,
#             "encoding"  : encoding,
#             "voter"     : {
#                 "name"        : matched["voter_name"],
#                 "voter_id"    : matched["voter_id"],
#                 "voter_number": matched.get("voter_number", ""),
#                 "voted_at"    : str(matched.get("voted_at", "")),
#             }
#         })

#     return jsonify({
#         "status"  : "new",
#         "message" : "New voter verified!",
#         "encoding": encoding
#     })


# @app.route("/api/cast-vote", methods=["POST"])
# def cast_vote():
#     """Save new voter to MongoDB."""
#     data         = request.json
#     name         = data.get("name", "").strip()
#     voter_number = data.get("voter_number", "").strip()
#     encoding     = data.get("encoding", [])

#     if not name or not encoding:
#         return jsonify({"success": False, "message": "Name and encoding required"}), 400

#     voter_id = generate_voter_id(encoding)
#     document = {
#         "voter_id"     : voter_id,
#         "voter_name"   : name,
#         "voter_number" : voter_number or f"VN-{voter_id[:8].upper()}",
#         "face_encoding": encoding,
#         "voted_at"     : datetime.utcnow(),
#         "status"       : "voted"
#     }
#     voters_col.insert_one(document)

#     return jsonify({
#         "success"     : True,
#         "message"     : "Vote cast successfully!",
#         "voter_id"    : voter_id,
#         "voter_name"  : name,
#         "voter_number": document["voter_number"],
#         "voted_at"    : str(document["voted_at"])
#     })


# @app.route("/api/stats", methods=["GET"])
# def get_stats():
#     """Dashboard statistics."""
#     total       = voters_col.count_documents({})
#     today       = datetime.utcnow().replace(hour=0, minute=0, second=0)
#     today_count = voters_col.count_documents({"voted_at": {"$gte": today}})

#     pipeline = [
#         {"$group": {"_id": {"$hour": "$voted_at"}, "count": {"$sum": 1}}},
#         {"$sort": {"_id": 1}}
#     ]
#     hourly      = list(voters_col.aggregate(pipeline))
#     hourly_data = [{"hour": f"{h['_id']}:00", "votes": h["count"]} for h in hourly]

#     return jsonify({
#         "total_voters": total,
#         "today_votes" : today_count,
#         "hourly_data" : hourly_data,
#     })


# @app.route("/api/voters", methods=["GET"])
# def get_voters():
#     """Return all voters without face encodings."""
#     voters = list(voters_col.find({}, {"_id": 0, "face_encoding": 0}))
#     for v in voters:
#         v["voted_at"] = str(v.get("voted_at", ""))
#     return jsonify({"voters": voters, "total": len(voters)})


# @app.route("/api/voters/<voter_id>", methods=["DELETE"])
# def delete_voter(voter_id):
#     """Delete a voter by ID."""
#     result = voters_col.delete_one({"voter_id": voter_id})
#     if result.deleted_count:
#         return jsonify({"success": True,  "message": "Voter deleted"})
#     return jsonify({"success": False, "message": "Voter not found"}), 404


# @app.route("/api/export/excel", methods=["GET"])
# def export_excel():
#     """Export voters to Excel."""
#     voters = list(voters_col.find({}, {"_id": 0, "face_encoding": 0}))
#     wb     = openpyxl.Workbook()
#     ws     = wb.active
#     ws.title = "Voters"

#     headers = ["#", "Name", "Voter Number", "Voter ID", "Voted At", "Status"]
#     for col, header in enumerate(headers, 1):
#         cell           = ws.cell(row=1, column=col, value=header)
#         cell.font      = Font(bold=True, color="FFFFFF")
#         cell.fill      = PatternFill("solid", fgColor="1a56db")
#         cell.alignment = Alignment(horizontal="center")

#     for i, voter in enumerate(voters, 1):
#         ws.append([
#             i,
#             voter.get("voter_name", ""),
#             voter.get("voter_number", ""),
#             voter.get("voter_id", ""),
#             str(voter.get("voted_at", "")),
#             voter.get("status", "voted")
#         ])

#     for col in ws.columns:
#         max_len = max(len(str(cell.value or "")) for cell in col)
#         ws.column_dimensions[col[0].column_letter].width = max_len + 4

#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)
#     return send_file(output,
#                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                      as_attachment=True,
#                      download_name="voters.xlsx")


# @app.route("/api/export/pdf", methods=["GET"])
# def export_pdf():
#     """Export voters to PDF."""
#     voters = list(voters_col.find({}, {"_id": 0, "face_encoding": 0}))
#     output = io.BytesIO()
#     doc    = SimpleDocTemplate(output, pagesize=letter)
#     styles = getSampleStyleSheet()
#     story  = []

#     story.append(Paragraph("Election Voters Report", styles["Title"]))
#     story.append(Paragraph(
#         f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC",
#         styles["Normal"]
#     ))

#     table_data = [["#", "Name", "Voter Number", "Voter ID", "Voted At"]]
#     for i, voter in enumerate(voters, 1):
#         table_data.append([
#             str(i),
#             voter.get("voter_name", ""),
#             voter.get("voter_number", ""),
#             voter.get("voter_id", "")[:12] + "...",
#             str(voter.get("voted_at", ""))[:16]
#         ])

#     table = Table(table_data, colWidths=[30, 120, 100, 110, 120])
#     table.setStyle(TableStyle([
#         ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1a56db")),
#         ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
#         ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
#         ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
#         ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
#         ("FONTSIZE",      (0, 0), (-1, -1), 9),
#     ]))
#     story.append(table)
#     doc.build(story)

#     output.seek(0)
#     return send_file(output,
#                      mimetype="application/pdf",
#                      as_attachment=True,
#                      download_name="voters.pdf")


# # ════════════════════════════════════════════════════════════
# if __name__ == "__main__":
#     print("\n╔══════════════════════════════════════════╗")
#     print("║   Election API Server Starting...        ║")
#     print("║   http://localhost:5000                  ║")
#     print("╚══════════════════════════════════════════╝\n")
#     app.run(debug=True, port=5000)